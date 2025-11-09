import openpyxl
from pathlib import Path

# مسار الملف
file_path = Path(r"C:\Users\haama\Desktop\PycharmProjects\Advanced Topics in Python\Excel\employees.xlsx")
wb = openpyxl.load_workbook(file_path)
sheet = wb["Employees"]

salaries = []

# قراءة الرواتب من العمود E (5)
for i in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=i, column=5).value  # العمود E
    try:
        salaries.append(float(cell_value))
    except:
        pass

# حساب المجموع وأعلى راتب
total_salary = sum(salaries)
max_salary = max(salaries)

# إضافة النتائج في الصف التالي
sheet.cell(row=sheet.max_row + 1, column=4).value = "Total Salary"
sheet.cell(row=sheet.max_row, column=5).value = total_salary

sheet.cell(row=sheet.max_row + 1, column=4).value = "Max Salary"
sheet.cell(row=sheet.max_row, column=5).value = max_salary

# حفظ الملف
wb.save(file_path)
print("Total and Max Salary added successfully!")
