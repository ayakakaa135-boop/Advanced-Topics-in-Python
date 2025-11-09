import openpyxl
from pathlib import Path

# إنشاء ملف جديد
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Students"

# إنشاء أوراق جديدة
workbook.create_sheet()
workbook.create_sheet(index=1, title='secondSheet')
workbook.create_sheet()
workbook.create_sheet(index=2, title='middleSheet')

# حذف ورقة معينة
del workbook['secondSheet']

# كتابة العناوين
sheet.append(["ID", "Name", "Age", "Grade", "Country"])

# بيانات تجريبية
data = [
    [1, "Liam", 20, 85, "Sweden"],
    [2, "Emma", 22, 92, "USA"],
    [3, "Noah", 21, 78, "UK"],
    [4, "Olivia", 23, 88, "Canada"],
    [5, "Ava", 20, 95, "Germany"],
]

# إضافة البيانات للورقة
for row in data:
    sheet.append(row)

# كتابة أسماء في العمود الثالث (C)
names = ['Hadi', 'Yara', 'Sara', 'Anas']

for i, name in enumerate(names, start=2):
    sheet.cell(row=i, column=3).value = name


# مسار الحفظ المطلوب
save_path = Path(r"C:\Users\haama\Desktop\PycharmProjects\Advanced Topics in Python\Excel\students_dataset.xlsx")

# حفظ الملف
workbook.save(save_path)

print("تم إنشاء الملف:", save_path)
