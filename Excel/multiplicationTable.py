import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Check that the user provided exactly one number as an argument
if len(sys.argv) == 2:
    try:
        number = int(sys.argv[1])
    except ValueError:
        print("❌ Error: Please enter a valid integer number.")
        sys.exit()

    print(f"\n📘 Creating a {number}x{number} multiplication table with nice formatting...\n")

    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Multiplication_Table"

    # Define styles
    bold_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Fill the table
    for row in range(number + 1):
        for col in range(number + 1):
            cell = sheet.cell(row=row + 1, column=col + 1)

            # Set cell value
            if row == 0 and col == 0:
                cell.value = ""
            elif col == 0:
                cell.value = row
                cell.font = bold_font
                cell.fill = header_fill
            elif row == 0:
                cell.value = col
                cell.font = bold_font
                cell.fill = header_fill
            else:
                cell.value = row * col

            # Apply alignment and borders
            cell.alignment = center_align
            cell.border = thin_border

    # Adjust column widths
    for column in sheet.columns:
        col_letter = column[0].column_letter
        sheet.column_dimensions[col_letter].width = 8

    # Define save path
    save_path = Path(r"C:\Users\haama\Desktop\PycharmProjects\Advanced Topics in Python\Excel")
    save_path.mkdir(parents=True, exist_ok=True)

    file_path = save_path / f"multiplication_table_{number}x{number}.xlsx"

    # Save the workbook
    workbook.save(file_path)

    print(f"✅ File created successfully at:\n{file_path}\n")

else:
    print("⚠️ Usage: python table.py <number>")
    print("Example: python table.py 10")
