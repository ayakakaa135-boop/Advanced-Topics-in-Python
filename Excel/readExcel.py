import openpyxl
from pathlib import Path

excelFile = openpyxl.load_workbook(
    Path.home() / Path('Desktop', 'PycharmProjects', 'Advanced Topics in Python', 'Excel', 'sample_data.xlsx')
)

print(excelFile.sheetnames)
sheet1 = excelFile['Sheet1']

print(sheet1.title)
activeSheet = excelFile.active
print(activeSheet.title)

print(sheet1['B1'].value)
print(sheet1['C2'].row)
print(sheet1['C2'].value)
print(sheet1['C3'].column)
print(sheet1['D1'].coordinate)
print(sheet1.cell(row=1, column=2).value)

print('-'*50)
for i in range(1, 7):
    print(i, sheet1.cell(row=i, column=1).value)

print('-'*50)

total = 0
for i in range(2, sheet1.max_row):
    name = sheet1.cell(row=i, column=2).value
    salary = sheet1.cell(row=i, column=5).value
    print(name, salary)
    total += salary

print('-'*50)
print("Total Salary =", total)
print(sheet1.max_row)
print(sheet1.max_column)
