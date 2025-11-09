from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path


number = 12


wb = Workbook()
ws = wb.active
ws.title = "Multiplication_Table"

# إعدادات التنسيق
bold_font = Font(bold=True, color="FFFFFF")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")


for row in range(number + 1):
    for col in range(number + 1):
        cell = ws.cell(row=row + 1, column=col + 1)

        if row == 0 and col == 0:
            cell.value = ""
        elif col == 0:
            cell.value = row
            cell.font = bold_font
            cell.fill = header_fill
        elif row == 0:
            cell.value = col
            cell.font = bold_font
            cell.fill = header_fill
        else:
            cell.value = row * col

        cell.alignment = center_align
        cell.border = thin_border


for column in ws.columns:
    col_letter = column[0].column_letter
    ws.column_dimensions[col_letter].width = 8


file_path = Path(__file__).parent / f"multiplication_table_{number}x{number}.xlsx"


wb.save(file_path)

print(f"✅ جدول الضرب {number}x{number} تم إنشاؤه وحفظه في:\n{file_path}")
